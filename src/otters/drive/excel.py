import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.layout import Layout, ManualLayout
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import column_index_from_string


def regression_to_excel(reg, name):

    # === Create workbook ===
    wb = Workbook()

    # --- Write Data to "Data" Sheet ---
    ws_chart = wb.active
    ws_chart.title = "Graphique"
    ws_data = wb.create_sheet(title="Données")

    reg.df.columns = pd.MultiIndex.from_tuples(tuple([tuple([name.replace('Model', 'Modèle') for name in col]) for col in reg.df.columns]))
    reg.df.columns = pd.MultiIndex.from_tuples(tuple([tuple([name.replace('Y_hat', 'Modèle') for name in col]) for col in reg.df.columns]))
    reg.df.columns = pd.MultiIndex.from_tuples(tuple([tuple([name.replace('X', 'Variables') for name in col]) for col in reg.df.columns]))
    reg.df.columns = pd.MultiIndex.from_tuples(tuple([tuple([name.replace('Y', 'Réelle') for name in col]) for col in reg.df.columns]))

    # I need to run this before deleting the intercept from the data
    # print(reg.df.columns.get_level_values(-1)[2:])
    # print(np.append(reg.reg.coef_[0], reg.reg.intercept_[0]))
    # print(np.append(reg.reg.p[0], 0))
    df_stats = pd.DataFrame({'Variable': reg.df.columns.get_level_values(-1)[2:], 'Coéfficient': np.append(reg.reg.coef_[0], reg.reg.intercept_[0]), 'Valeurs P': np.append(reg.reg.p[0], 0)})

    reg.df.drop(columns=[("intercept", "intercept")], inplace=True)
    # print(reg.df.columns)
    # break

    reg.df.columns.set_names(["Type", "Date"], level=[0, 1], inplace=True)

    for r in dataframe_to_rows(reg.df, index=True, header=True):
        ws_data.append(r)

    ws_data['A2'] = "Date"
    ws_data.delete_rows(3)

    for row in range(3, reg.df.shape[0] + 1):
        cell = ws_data[f"A{row}"]
        cell.number_format = 'yyyy-mm-dd'  # or another Excel-compatible format

    # # --- Create Chart Sheet ---
    # # Create line chart
    chart = LineChart()
    chart.title = name
    chart.style = 2
    chart.y_axis.title = ''
    chart.x_axis.title = ''
    chart.width = 30
    chart.height = 13.5

    chart.x_axis = chart.x_axis  
    chart.y_axis = chart.y_axis
    chart.x_axis.delete = False
    chart.y_axis.delete = False

    # Reference data
    data = Reference(ws_data, min_col=2, max_col=3, min_row=2, max_row=reg.df.shape[0])  # Y column
    cats = Reference(ws_data, min_col=1,  min_row=3, max_row=reg.df.shape[0])  # X column (categories)

    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)

    # Optional: control position/size
    # chart.layout = Layout(
    #     manualLayout=ManualLayout(
    #         # x=0.25, y=0.1,
    #         h=0.6, w=0.6
    #     )
    # )

    # Add chart to "Chart" tab at a specific cell
    ws_chart.add_chart(chart, "B5")

    ws_chart['C31'] = reg.equation


    # Add stats to "Chart" tab
    stats = {
        # "Intercept": reg.reg.intercept_,
        # "Slope": slope,
        "R²": reg.score,
        # "Mean Y": df['Y'].mean(),
        # "Std Y": df['Y'].std()

    }
    stats = {
        'début' : reg.start,
        'fin' : reg.end,
        'R2': reg.score,
        # 'CV (low confidence)': reg.cv,
        'CV-RMSE': f"{reg.cvrmse:.1%}",
        # 'P-Values': reg.reg.p,
        # 'N_Samples': reg.Y.shape[0],
    }
    ws_chart[f'T4'] = "Statistique"
    ws_chart[f'U4'] = "Valeur"
    for i, item in enumerate(stats.items()):
        if type(item[1]) == list:
            value = ", ".join([str(val) for val in item[1]])
        elif type(item[1]) == np.float64:
            value = '{:.4F}'.format(item[1])
        else:
            value = item[1]
        

        ws_chart[f'T{i+5}'] = item[0]
        ws_chart[f'U{i+5}'] = value

    # Format the tables
    # Define the table range (must include headers)
    table_range = f"T4:U8"

    # Create the table
    table = Table(displayName="stats", ref=table_range)

    # Apply a style
    style = TableStyleInfo(
        name="TableStyleMedium9",  # any built-in Excel style
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    table.tableStyleInfo = style
    ws_chart.add_table(table)

    # print(reg.df.columns.get_level_values(-1)[2:])
    # print(np.append(reg.reg.coef_[0], reg.reg.intercept_[0]))



    # ws_stats = wb.create_sheet(title="Stats")

    # for r in dataframe_to_rows(df_stats, index=False, header=True):
    #     ws_chart.rows(r)
    # ws_stats.delete_rows(2)

    start_row = 10
    start_col_letter = "T"
    start_col = column_index_from_string(start_col_letter)

    # Paste DataFrame starting at C5
    for r_idx, row in enumerate(dataframe_to_rows(df_stats, index=False, header=True), start=start_row):
        for c_idx, value in enumerate(row, start=start_col):
            ws_chart.cell(row=r_idx, column=c_idx, value=value)

    wb.save(f"model_outputs/{name}.xlsx",)

    return