"""
The file_loader module provides functions used to load data from files (excel, csv, pdf, etc) or into files.
"""
import pandas as pd
import yaml
import os
from glob import glob
import inspect
import warnings
import re

from .wrangler import two_letter_month_to_number
from .time_tools import selectiveResample

import pymupdf
import tabula
import openpyxl
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from openpyxl.utils import get_column_letter
os.environ["XLWINGS_LICENSE_KEY"] = "noncommercial"

def replace_backslashes_in_dict(dictionary):
    for key, val in dictionary.items():
        if isinstance(val, list):
            # dictionary.update({key: replace_backslashes_in_dict(val)})
            dictionary.update({key: [term.replace('\\', "/") for term in val]})
        elif isinstance(val, dict):
            dictionary.update({key: replace_backslashes_in_dict(val)})
        elif isinstance(val, str):
            dictionary.update({key: val.replace('\\', "/")})
    return dictionary

def import_config(configFolder='', recursive=True):
    """
    MUST USE SINGLE QUOTES IN YAML FILES FOR BACKSLASHES TO WORK
    Import all config files from the supplied folder. Defaults to the root folder  
    Any files ending in 'config.yaml' or 'config.xlsx' will be treated. So you could have a formatting config called 'format.config.yaml'  
    
    **Parameters:**
    > **configFolder:** *string, default: empty*  
    >> The relative path to the config folder. 

    > **recursive:** *boolean, default: `True`* 
    >> Returns all files in dir and sub-dirs if `True`. Only files in root dir if `False`.

    **Returns:**  
    > **dict*
    """

    # Load any config files
    if recursive == True:
        yamlFiles = glob(os.getcwd()+configFolder+'/**/*config.yaml', recursive=True)
        xlFiles = glob(os.getcwd()+configFolder+'/**/*config.xlsx', recursive=True)
    else:
        yamlFiles = glob(os.getcwd()+configFolder+'/./*config.yaml', recursive=False)
        xlFiles = glob(os.getcwd()+configFolder+'/./*config.xlsx', recursive=True)

    yamlFiles = [file.replace(os.getcwd(), '').lstrip('/\\') for file in yamlFiles]
    xlFiles = [file.replace(os.getcwd(), '').lstrip('/\\') for file in xlFiles]
    xlFiles = [file for file in xlFiles if '~' not in file]

    if yamlFiles or xlFiles:
        print(f'Loading config files: {yamlFiles+xlFiles}')
    else:
        configFile = 'config.yaml'
        func_dir_path = '/'.join(os.path.abspath(inspect.getfile(import_config)).split('\\')[0:-1])
        print(f"""There doesn't appear to be a config file in the supplied folder. No config was imported""")

    config = {}
    for file in yamlFiles:
        print(f"file: {file}")
        with open(file, encoding='utf8') as f:
            try:
                config.update(yaml.load(f, Loader=yaml.FullLoader))
            except:
                raise Exception("That didn't work, the bug shown above is likely a bug in your config file")
            f.close()

    for file in xlFiles:
        df = pd.read_excel(file)
        dictName = df.columns[0]
        df.columns = df.iloc[0 ,:]
        df = df.iloc[1: ,:]
        config.update({dictName: df.T.to_dict()})

    config = replace_backslashes_in_dict(config)

    return config

def getFiles(path='*', recursive=False, fileTypes=['']):
    """
    Get list of all files in a supplied path. Can pass file types  

    **Parameters:**  
    > **path: *string, default: empty***  
    >> Pass a relative or absolute path

    > **recursive: *bool, default: False***  
    >> If True search recursively, meaning it will search the folder supplied and any subfolders.

    > **fileTypes: *list, default: empty list***  
    >> Example: ['.xlsx', '.xls']

    <a href="https://www.youtube.com/watch?v=dQw4w9WgXcQ">Click here for more info</a>

    **Returns:**  
    > **list**  
    """
    files = []
    for fileType in fileTypes:
        files.extend(glob(path+fileType, recursive=recursive))

    return list(set(files))

def getExcelDFs(path='*', recursive=False, verbose=False,):
    """
    Import all Excel files from a path into a list of DataFrames  

    **Parameters:**
    > **path: *string, default: "/*"***  
    >> Pass a relative or absolute path

    > **recursive: *bool, default: False***  
    >> If True search recursively, meaning it will search the folder supplied and any subfolders.

    > **verbose: *bool, default: False***  
    >> If True prints each file as it's being imported.

    **Returns:**  
    > **list of DataFrames**  
    """
    fileTypes = ['.xlsx', '.xlsm' , '.xls',]
    files = getFiles(path=path, recursive=recursive, fileTypes=fileTypes)

    if verbose:
        print(f'File list:\n{files}\n')

    dfList = []
    for file in files:
        if verbose:
            print(f'Importing file: {file}')
        chunk = pd.read_excel(file,)
        dfList.append(chunk)
    return dfList

def formatData(wb_name, sheet='Sheet1'):
    """
    This will format an outputted excel file so it doesn't make your eyes bleed.  
    Most useful for outputting DataFrames to Excel.  
    Stole it from StackOverflow. It's not super optimal but it's concise which I like.

    **Parameters:**  
    > **wb_name: *str, Required***  
    >> The name of the file. Can be relative and prolly absolute

    > **sheet: *string, default: "Sheet1"*** 
    >> The name of the sheet. Leave blank if you didn't rename the sheet and it's the first/only one in the workbook.

    **Returns:**  
    > **None**
    """
    wb = openpyxl.load_workbook(wb_name)
    ws = wb[sheet]
    
    dim_holder = DimensionHolder(worksheet=ws)
    
    for col in range(ws.min_column, ws.max_column + 1):
        dim_holder[get_column_letter(col)] = ColumnDimension(ws, min=col, max=col, width=19)
    
    ws.column_dimensions = dim_holder

    for row in ws.iter_rows():
        for cell in row:      
            cell.alignment =  cell.alignment.copy(wrapText=True)
        break # Oof lol, a break in production. Somebody bonk this guy

    wb.save(wb_name)
    wb.close()
    return

def save2xl(df, file='', sheet='Data', startCell=[3, 2], table=True, visible=False):
    import xlwings as xw
    """
    IN DEVELOPMENT, PROBABLY WILL NOT WORK  

    Saves a DataFrame to an excel file  
     
    It will either overwrite the sheet in an existing file or create a new one  

    Accepts relative or absolute paths  
    """
    if glob(file):
        xw.Book(file).sheets[sheet]
    else:
        
        xw.Book().sheets[sheet]
    
    #     Clear Filters
    if sheet.api.AutoFilter:
        sheet.api.AutoFilter.ShowAllData()

    #get the number of columns/rows the query has
    queryCols, queryRows = createDF(sheet).shape
    print(queryCols)
#     queryRows = createDF(sheet).shape[0]
    #if queryCols is less than the # of columns the Python makes, we'll insert those columns as blanks and then paset the Python overtop.
    if (df.shape[1] > queryCols):
        for col in range(queryCols+2, df.shape[1]+2):
            sheet.api.Columns(queryCols+2).Insert()
    
    #Paste the treated DataFrame into excel
    sheet.range(startCell).expand('table').options(index=True, header=True, empty="0").value = df  

    sht = xw.Book(fileOut).sheets['Data']
    # Put an if here being like: If it's a timestamp then format it nicely as a timeseries
    return

def standardize_num_format(df):
    """Undocumented
    Source: C:/Users/LucaLafontaine/AKONOVIA/EMO - Documents/002-ALCOVI/22-673 VSL - Ajustement Lufa/1-Intrant/Factures/Énergir/read_JCI_data.ipynb
    THis is used internally and just for the invoice reading (energir) functions
    """
    # print(df.dtypes)
    df = df.copy()
    
    for col in [col for col in df.columns if df[col].dtypes == object]:
        df.loc[:, col] = df[col].str.replace(' ', '')
        df.loc[:, col] = df[col].str.replace(',', '.')
        df.loc[:, col] = df[col].str.replace('Réelle', '')

        for symbol in ['%', r"\$", '¢', "°C"]:
            if df[col].str.contains(symbol).any():
                # print(f"{symbol} {col}")
                # print(df[col].str.contains(symbol, regex=True))
                df.loc[:, col] = df[col].str.replace(symbol, '', regex=True)
                new_col = col+" {}".format(symbol.replace(r'\\', ''))
                df.rename(columns={col:new_col}, inplace=True)
                col = new_col
                # print(col)
        # print(df.columns)
        # print(df[col])
        try:
            df[col] = pd.to_numeric(df[col], )
        except:
            pass
    return df

def get_table_gaz(file_name, resample=False):
    """Undocumented
    Source: C:/Users/LucaLafontaine/AKONOVIA/EMO - Documents/002-ALCOVI/22-673 VSL - Ajustement Lufa/1-Intrant/Factures/Énergir/read_JCI_data.ipynb
    Untested in this environment. Want to se if having it here will make development faster
    """
    doc = pymupdf.open(file_name)
    number_pages = doc.page_count
    for i in range(number_pages):
        page = doc[i]

        ## Find METER NUMBER
        meter_locations = page.search_for("Numéro de compte")
        if len(meter_locations)>0:
            ## Take only the first meter 
            meter_location = meter_locations[0]

            top = meter_location[1]
            left = meter_location[0]
            bottom = meter_location[3]+40
            right = meter_location[2]+20
            meter_number_list = tabula.read_pdf(file_name, lattice=False,stream = True,  pages=i+1, area=(top, left, bottom, right))
            meter_number = int(meter_number_list[0].iloc[0,0].replace(' ', ''))
        else: 
            warnings.warn("No meter was found for the file:\n {file_name}\nContinuing with meter number: 0")
            meter_number = 0

        ## Find Invoice number so that the invoices can be sorted
        invoice_locations = page.search_for("Facture n°")
        if len(invoice_locations)>0:
            ## Take only the first invoice number location
            invoice_location = invoice_locations[0]
            top = invoice_location[1]-5
            left = invoice_location[0]+25
            bottom = invoice_location[3]+5
            right = invoice_location[2]+100
            invoice_number_list = tabula.read_pdf(file_name, lattice=False,stream = True,  pages=i+1, area=(top, left, bottom, right))
            invoice_number_df = invoice_number_list[0]
            invoice_number = int(invoice_number_df.columns[1].replace(' ', '')) # We're taking the second header name... i don't love this
        else: 
            warnings.warn("No invoice number was found for the file:\n {file_name}\nContinuing with invoice number: 0")
            invoice_number = 0
    
        tables_historique = page.search_for("Historique de consommation")
        if len(tables_historique)==0:
            tables_historique = page.search_for("Consommations antérieures")

        if len(tables_historique)>0:
            table_historique = tables_historique[0]

            ## SHift the box down to get the headers of the actual table
            top = table_historique[1]+20
            left = table_historique[0]-20
            bottom = table_historique[3]+30
            right = table_historique[2]+95
            headers = tabula.read_pdf(file_name, lattice=False,stream = True,  pages=i+1, area=(top, left, bottom, right))

            header = headers[0]
            columns = list(zip(header.columns, header.iloc[0,:]))
            columns = [(' '.join(str(i) for i in x)) for x in columns]
            columns = [col.replace('Unnamed: 0', '').strip() for col in columns]
            header = columns
            top = bottom-10
            left = left
            bottom = bottom+120
            right = right
            table = tabula.read_pdf(file_name, lattice=False,stream = True,  pages=i+1,area=(top, left, bottom, right))[0].fillna("")
            table.columns = header

            table['Début'] = table.apply(lambda x: (' ').join(x["Période du"].split(' ')[:3]), axis=1).astype('unicode')
            table['Début'] = pd.to_datetime(table['Début'].apply(lambda x: two_letter_month_to_number(x)), format='%d %m %Y')
            table['Fin'] = table.apply(lambda x: (' ').join(x["au"].split(' ')[:3]), axis=1).astype('unicode')
            table['Fin'] = pd.to_datetime(table['Fin'].apply(lambda x: two_letter_month_to_number(x)), format='%d %m %Y')

            table['Volume (m3)'] = table['Volume (m3)'].str.replace(r"[A-Z\s]", "", regex=True).astype("float")
            table = standardize_num_format(table)
            table = table.drop(['Période du', "au", ], axis=1)
            table.index = table['Début']
            
            table = table.loc[:, ["Fin", "Volume (m3)", "Montant* ($)"]] 

            table['invoice_number'] = invoice_number
            table['meter_number'] = meter_number
        
            if resample:
                tableResampled = table.reset_index()
                tableResampled = pd.concat([tableResampled.set_index('Début'), tableResampled.set_index('Fin')]).drop(["Début", "Fin"], axis=1)
                tableResampled = tableResampled.resample("D").mean().ffill()
                tableResampled['Jours'] = 1
                tableResampled = selectiveResample(tableResampled, 'MS', tableResampled.columns.drop("Jours"), "Jours") 

                return tableResampled
            else:
                return table
        
def get_table_tarif_M(file_name, resample=False):
    """Undocumented
    Source: C:/Users/LucaLafontaine/AKONOVIA/EMO - Documents/002-ALCOVI/22-673 VSL - Ajustement Lufa/1-Intrant/Factures/Énergir/read_JCI_data.ipynb
    Untested in this environment. Want to se if having it here will make development faster
    """
    doc = pymupdf.open(file_name)
    number_pages = doc.page_count
    for i in range(number_pages):
        page = doc[i]

        ## Find METER NUMBER
        meter_locations = page.search_for("Numéro de compte")
        if len(meter_locations)>0:
            ## Take only the first meter number location
            meter_location = meter_locations[0]

            top = meter_location[1]
            left = meter_location[0]
            bottom = meter_location[3]+25
            right = meter_location[2]+33
            meter_number_list = tabula.read_pdf(file_name, lattice=False,stream = True,  pages=i+1, area=(top, left, bottom, right))
            meter_number_df = meter_number_list[0]
            meter_number = int(meter_number_df.iloc[0,0].replace(' ', ''))
        else: 
            warnings.warn("No meter was found for the file:\n {file_name}\nContinuing with meter number: 0")
            meter_number = 0

        ## Find Invoice number so that the invoices can be sorted
        invoice_locations = page.search_for("Numéro de facture")
        if len(invoice_locations)>0:
            ## Take only the first invoice number location
            invoice_location = invoice_locations[0]

            top = invoice_location[1]
            left = invoice_location[0]
            bottom = invoice_location[3]+25
            right = invoice_location[2]+33
            invoice_number_list = tabula.read_pdf(file_name, lattice=False,stream = True,  pages=i+1, area=(top, left, bottom, right))
            invoice_number_df = invoice_number_list[0]
            invoice_number = int(invoice_number_df.iloc[0,0].replace(' ', ''))
        else: 
            warnings.warn("No invoice number was found for the file:\n {file_name}\nContinuing with invoice number: 0")
            invoice_number = 0
    
        tables_historique = page.search_for("HISTORIQUE DE LA CONSOMMATION D’ÉLECTRICITÉ")
        if len(tables_historique)==0:
            tables_historique = page.search_for("Consommations antérieures")

        if len(tables_historique)>0:
            table_historique = tables_historique[0]

            ## SHift the box down to get the headers of the actual table
            top = table_historique[1]+10
            left = table_historique[0]-20
            bottom = table_historique[3]+30
            right = table_historique[2]+150
            headers = tabula.read_pdf(file_name, lattice=False,stream = True,  pages=i+1, area=(top, left, bottom, right))

            header = headers[0]
            columns = list(zip(header.columns, header.iloc[0,:]))
            columns = [(' '.join(str(i) for i in x)) for x in columns]
            columns = [re.sub('Unnamed: .+? ', '', col).strip() for col in columns]
            header = columns
            top = bottom-20
            left = left
            bottom = bottom+105
            right = right
            table = tabula.read_pdf(file_name, lattice=False,stream = True,  pages=i+1,area=(top, left, bottom, right))[0].fillna("")
            table.columns = header

            table['Début'] = table.apply(lambda x: (' ').join(x["Du"].split(' ')[:3]), axis=1).astype('unicode')
            table['Fin'] = table.apply(lambda x: (' ').join(x["Au"].split(' ')[:3]), axis=1).astype('unicode')
            table = standardize_num_format(table, )
            table.index = table['Début']
            
            table = table.loc[:, [
                "Fin", 'kWh', 'Moyenne (kWh/j)',
                'Puissance facturée (kW)', 'Temp. ext. moyenne °C','Montant (taxes comprises) $'
                ]
            ]
            table['invoice_number'] = invoice_number
            table['meter_number'] = meter_number

            if resample:
                tableResampled = table.reset_index()

                tableResampled = pd.concat([tableResampled.set_index('Début'), tableResampled.set_index('Fin')]).drop(["Début", "Fin"], axis=1)
                tableResampled.index = pd.to_datetime(tableResampled.index)
                tableResampled.index.name = 'Timestamp'

                tableResampled = tableResampled.resample("D").mean().ffill()
                tableResampled['Jours'] = 1
                tableResampled = selectiveResample(tableResampled, 'MS', tableResampled.columns.drop("Jours"), "Jours") 

                return tableResampled
            else:
                return table
            
def clean_tableau_des_factures_file(file, facture_type_col = 'Fournisseur', time_cols = ['De', 'À'], format="%d/%m/%Y %H:%M:%S", verbose=False):
    """
    Method to sort the bills that come out of JOOL (Tested on Kelvin, not Pub), set them to timestamps, and split the bills by fournisseur.

    Puts this all back into the same file under new sheets for each fournisseur 
    """
    if verbose:
        print(f"fixing time columns and splitting out bills by fournisseur then sorting for file:  \n{file}")
    
    df = pd.read_excel(file)
    for col in time_cols:
        df[col] = pd.to_datetime(df[col], format=format)

    for bill_type in df[facture_type_col].unique():
        df_bill = df.loc[df[facture_type_col] == bill_type, :]
        df_bill = df_bill.sort_values(time_cols[0])
   
        with pd.ExcelWriter(file, mode='a', engine='openpyxl', if_sheet_exists='replace') as writer:
            df_bill.to_excel(writer, sheet_name=bill_type)
        
        formatData(file, bill_type)

    if verbose:
        print('Done')
    
    return